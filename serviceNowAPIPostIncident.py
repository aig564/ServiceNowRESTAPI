#Need to install requests package for python
#sudo easy_install requests
import requests
import json
import getpass 
import openpyxl

# Set the request parameters
url = 'https://conns.service-now.com/api/now/table/incident'# ?sysparm_query=assigned_to.nameLIKEAlfonzo^stateIN6,7^resolved_atBETWEENjavascript:gs.beginningOfLast7Days()@javascript:gs.endOfCurrentHour()^short_descriptionLIKEvary&sysparm_limit=2'# &sysparm_fields=number,description,short_description' # &sysparm_limit=10'

user = getpass.getuser()   

try: 
    pwd = getpass.getpass("User Name : %s\nPassword: " % user)  
except Exception as error: 
    print('ERROR', error) 

# endUserID = input('Enter UserName to search: \n')
assignedToUser = input('Enter your employee #:\n')
while assignedToUser.isdigit()==False or len(assignedToUser)!=5:
	try:
		assignedToUser = input('Enter your employee #:\n')
	except: # just catch the exceptions you know!
		continue 

headers = {"Accept":"application/json"}
newUrl = 'https://conns.service-now.com/api/now/table/sys_user?sysparm_query=employee_numberISNOTEMPTY^employee_number=' + assignedToUser + '&sysparm_fields=name&sysparm_exclude_reference_link=TRUE&sysparm_display_value=TRUE'
response = requests.get(newUrl, auth=(user, pwd), headers=headers )
if response.status_code != 200: 
    print('Status:', response.status_code, 'Headers:', response.headers, 'Error Response:',response.json())
    exit()
serviceNowData = response.json()
users = serviceNowData['result']
assignedToUser = users[0]['name']
print(assignedToUser)

caller = input('Enter employee # of Caller (please include leading zeros - ID should be 5 digits long): \n')
headers = {"Accept":"application/json"}
newUrl = 'https://conns.service-now.com/api/now/table/sys_user?sysparm_query=employee_numberISNOTEMPTY^employee_number=' + caller + '&sysparm_fields=name&sysparm_exclude_reference_link=TRUE&sysparm_display_value=TRUE'
response = requests.get(newUrl, auth=(user, pwd), headers=headers )
if response.status_code != 200: 
    print('Status:', response.status_code, 'Headers:', response.headers, 'Error Response:',response.json())
    exit()
serviceNowData = response.json()
users = serviceNowData['result']
caller = users[0]['name']
print(caller)

ticketDetail = input('Enter detail of ticket : \n')

# option = int(input('Enter option for type of ticket to be created : \n1. Session Reset\n2. Password Reset\n3. Account Unlock\n4. Invoice Reprint\n'))

option = 0
while option not in range(1, 4):
	try:
		option = int(input('Enter option for type of ticket to be created : \n1. Session Reset\n2. Password Reset\n3. Account Unlock\n4. Invoice Reprint\n'))
	except: # just catch the exceptions you know!
		continue # or print a message such as print("Bad choice. Try again...")
if(int(option)==1):
	dataUpdate = '<request><entry><active>True</active><assigned_to>'+assignedToUser+'</assigned_to><assignment_group>IT Helpdesk - San Antonio</assignment_group><caller_id>'+caller+'</caller_id><category>AS/400</category><child_incidents>0</child_incidents><close_code>Solved (Permanently)</close_code><contact_type>Phone</contact_type><description>User requested to vary on/reset session.</description><escalation>Normal</escalation><impact>1 - Company Wide / Site</impact><incident_state>Resolved</incident_state><notify>Do Not Notify</notify><priority>4 - Low</priority><severity>3 - Low</severity><short_description>User requested session reset or to vary on terminal. - '+ticketDetail+'</short_description><state>Resolved</state><subcategory>Session reset / Varied off</subcategory><u_work_type>KTLO</u_work_type></entry></request>'
if(int(option)==2):
	dataUpdate = '<request><entry><active>True</active><assigned_to>'+assignedToUser+'</assigned_to><assignment_group>IT Helpdesk - San Antonio</assignment_group><caller_id>'+caller+'</caller_id><category>Password Reset</category><child_incidents>0</child_incidents><close_code>Solved (Permanently)</close_code><contact_type>Phone</contact_type><description>User requested a password reset for</description><escalation>Normal</escalation><impact>3 - Single User</impact><incident_state>Resolved</incident_state><notify>Do Not Notify</notify><priority>4 - Low</priority><severity>3 - Low</severity><short_description>User requested a password reset for '+ticketDetail+'</short_description><state>Resolved</state><subcategory>Other</subcategory><u_work_type>KTLO</u_work_type></entry></request>'
	
if(int(option)==3):
	dataUpdate = '<request><entry><active>True</active><assigned_to>'+assignedToUser+'</assigned_to><assignment_group>IT Helpdesk - San Antonio</assignment_group><caller_id>'+caller+'</caller_id><category>Access</category><child_incidents>0</child_incidents><close_code>Solved (Permanently)</close_code><contact_type>Phone</contact_type><description>User requested an account be unlocked.</description><escalation>Normal</escalation><impact>3 - Single User</impact><incident_state>Resolved</incident_state><notify>Do Not Notify</notify><priority>4 - Low</priority><severity>3 - Low</severity><short_description>User requested an account be unlocked for '+ticketDetail+'</short_description><state>Resolved</state><subcategory>Account Lockout</subcategory><u_work_type>KTLO</u_work_type></entry></request>'
if(int(option)==4):
	dataUpdate = '<request><entry><active>True</active><assigned_to>'+assignedToUser+'</assigned_to><assignment_group>IT Helpdesk - San Antonio</assignment_group><caller_id>'+caller+'</caller_id><category>Printer</category><child_incidents>0</child_incidents><close_code>Solved (Permanently)</close_code><contact_type>Phone</contact_type><description>User requested reprint in as400</description><escalation>Normal</escalation><impact>1 - Company Wide / Site</impact><incident_state>Resolved</incident_state><notify>Do Not Notify</notify><priority>4 - Low</priority><severity>3 - Low</severity><short_description>User requested a reprint for invoice ending in '+ticketDetail+'</short_description><state>Resolved</state><subcategory>Reprint Request</subcategory><u_work_type>KTLO</u_work_type></entry></request>'
print(dataUpdate)



# if(int(option)==1):
	# dataUpdate = '<request><entry><active>True</active><assigned_to>'+assignedToUser+'</assigned_to><assignment_group>IT Helpdesk - San Antonio</assignment_group><caller_id>'+caller+'</caller_id><category>AS/400</category><child_incidents>0</child_incidents><close_code>Solved (Permanently)</close_code><contact_type>Phone</contact_type><description>User requested to vary on/reset session. '+ticketDetail+'</description><escalation>Normal</escalation><impact>1 - Company Wide / Site</impact><incident_state>Resolved</incident_state><notify>Do Not Notify</notify><priority>4 - Low</priority><severity>3 - Low</severity><short_description>User requested session reset or to vary on terminal. - WH80PC0110</short_description><state>Resolved</state><subcategory>Session reset / Varied off</subcategory><u_work_type>KTLO</u_work_type></entry></request>'
# if(int(option)==2):
	# dataUpdate = '<request><entry><active>True</active><assigned_to>'+assignedToUser+'</assigned_to><assignment_group>IT Helpdesk - San Antonio</assignment_group><caller_id>'+caller+'</caller_id><category>Password Reset</category><child_incidents>0</child_incidents><close_code>Solved (Permanently)</close_code><contact_type>Phone</contact_type><description>User requested a password reset for</description><escalation>Normal</escalation><impact>3 - Single User</impact><incident_state>Resolved</incident_state><notify>Do Not Notify</notify><priority>4 - Low</priority><severity>3 - Low</severity><short_description>User requested a password reset for '+ticketDetail+'</short_description><state>Resolved</state><subcategory>Other</subcategory><u_work_type>KTLO</u_work_type></entry></request>'
	
# if(int(option)==3):
	# dataUpdate = '<request><entry><active>True</active><assigned_to>'+assignedToUser+'</assigned_to><assignment_group>IT Helpdesk - San Antonio</assignment_group><caller_id>'+caller+'</caller_id><category>Access</category><child_incidents>0</child_incidents><close_code>Solved (Permanently)</close_code><contact_type>Phone</contact_type><description>User requested an account be unlocked.</description><escalation>Normal</escalation><impact>3 - Single User</impact><incident_state>Resolved</incident_state><notify>Do Not Notify</notify><priority>4 - Low</priority><severity>3 - Low</severity><short_description>User requested an account be unlocked for '+ticketDetail+'</short_description><state>Resolved</state><subcategory>Account Lockout</subcategory><u_work_type>KTLO</u_work_type></entry></request>'
# if(int(option)==4):
	# dataUpdate = '<request><entry><active>True</active><assigned_to>'+assignedToUser+'</assigned_to><assignment_group>IT Helpdesk - San Antonio</assignment_group><caller_id>'+caller+'</caller_id><category>Printer</category><child_incidents>0</child_incidents><close_code>Solved (Permanently)</close_code><contact_type>Phone</contact_type><description>User requested reprint in as400</description><escalation>Normal</escalation><impact>1 - Company Wide / Site</impact><incident_state>Resolved</incident_state><notify>Do Not Notify</notify><priority>4 - Low</priority><severity>3 - Low</severity><short_description>User requested a reprint for invoice ending in '+ticketDetail+'</short_description><state>Resolved</state><subcategory>Reprint Request</subcategory><u_work_type>KTLO</u_work_type></entry></request>'
	# wb = openpyxl.load_workbook('IncidentPost.xlsx')
	# sheet = wb.active
	# variableNames = [None]*sheet.max_column
	# variableUpdates = [None]*sheet.max_column

	# for column in range(1,sheet.max_column):
		# variableNames[column] = str(sheet.cell(row=1,column=column).value)
	
	# for column in range(1,sheet.max_column):
		# variableUpdates[column]=sheet.cell(row=int(option)+1, column=column).value
	# dataUpdate='<request><entry>'
	# for column in range (1,sheet.max_column-2):
		# dataUpdate = dataUpdate + '<' + str(variableNames[column]) + '>' + str(variableUpdates[column]) + '</' + str(variableNames[column]) +'>'
	
	# dataUpdate = dataUpdate + '<' + str(variableNames[sheet.max_column-1]) + '>' + str(variableUpdates[sheet.max_column-1]) + '</' + str(variableNames[sheet.max_column-1]) +'>' + '</entry></request>'

	# textFile = open('IncidentPostData.txt', 'w')
	# textFile.write('Incident Post Data String: \n\n')
	# textFile.write(dataUpdate)
	# textFile.close()
	# print(dataUpdate)
	# exit()
# else:
# 	option = input('Enter option for type of ticket to be created : \n1. Session Reset\n2. Password Reset\n3. Account Unlock\n4. Invoice Reprint\n')

# print(str(userNumber+1)+': '+users[userNumber]['user_name']+'('+users[userNumber]['sys_id']+')'+'('+users[userNumber]['manager']+')')
# dataUpdate='{"'
# for column in range (1,sheet.max_column-2):
	# dataUpdate = dataUpdate + str(variableNames[column]) + '":"' + str(variableUpdates[column]) + '","'
	
# dataUpdate = dataUpdate + str(variableNames[sheet.max_column-1]) + '":"' + str(variableUpdates[sheet.max_column-1]) + '"}'

# dataUpdate='<request><entry>'
# for column in range (1,sheet.max_column-2):
	# dataUpdate = dataUpdate + '<' + str(variableNames[column]) + '>' + str(variableUpdates[column]) + '</' + str(variableNames[column]) +'>'
	
# dataUpdate = dataUpdate + '<' + str(variableNames[sheet.max_column-1]) + '>' + str(variableUpdates[sheet.max_column-1]) + '</' + str(variableNames[sheet.max_column-1]) +'>' + '</entry></request>'

# textFile = open('IncidentPostData.txt', 'w')
# textFile.write('Incident Post Data String: \n\n')
# textFile.write(dataUpdate)
# textFile.close()
# print(dataUpdate)
# exit()

 
# Set proper headers
headers = {"Content-Type":"application/xml","Accept":"application/json"}

# dataUpdate = '<request><entry><active>False</active><assigned_to>Alfonzo Galvan</assigned_to><assignment_group>IT Helpdesk - San Antonio</assignment_group><caller_id>' + caller + '</caller_id><category>AS/400</category><child_incidents>0</child_incidents><close_code>Solved (Permanently)</close_code><contact_type>Phone</contact_type><description>User requested to vary on/reset session.</description><escalation>Normal</escalation><impact>3 - Single User</impact><incident_state>Closed</incident_state><notify>Do Not Notify</notify><priority>4 - Low</priority><severity>3 - Low</severity><short_description>User requested session reset or to vary on terminal. - ' + ticketDetail + '</short_description><state>Closed</state><subcategory>Session reset / Varied off</subcategory><u_work_type>KTLO</u_work_type></entry></request>'

# Do the HTTP request
exit()
response = requests.post(url, auth=(user, pwd), headers=headers, data=dataUpdate)
 
# Check for HTTP codes other than 200
if response.status_code != 200: 
    print('Status:', response.status_code, 'Headers:', response.headers, 'Error Response:',response.json())
    exit()
 
# Decode the JSON response into a dictionary and use the data
# print('Status:',response.status_code,'Headers:',response.headers,'Response:',response.json())

serviceNowData = response.json()

print(type(serviceNowData))
print(serviceNowData)

# tickets = serviceNowData['result']

# print(type(tickets))
# print(tickets)
# for ticketNumber in tickets:

# print('Cookies', response.cookies)
