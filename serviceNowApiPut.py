#Need to install requests package for python
#sudo easy_install requests
import requests
import openpyxl
import getpass 
  
user = getpass.getuser()   

try: 
    pwd = getpass.getpass("User Name : %s\nPassword: " % user)  
except Exception as error: 
    print('ERROR', error) 
# else: 
    # print('Password entered:', pwd) 
	


# Set the request parameters
url = 'https://conns.service-now.com/api/now/table/cmdb_ci_computer/'


# Set proper headers
headers = {"Content-Type":"application/json","Accept":"application/json"}

# Open Excel Workbook

wb = openpyxl.load_workbook('SATXDTHOMEPCs-REMOVEDESKLOCATION.xlsx')
sheet = wb.active

# Iterate through Workbook

variableNames = [None]*sheet.max_column
variableUpdates = [None]*sheet.max_column

for column in range(sheet.max_column,1,-1):
	variableNames[sheet.max_column-column] = str(sheet.cell(row=1,column=column).value)
	

for row in range(2, sheet.max_row+1):
	# hardwareStatus = sheet.cell(row=row, column=1).value
	# installStatus = 7# sheet.cell(row=row, column=2).value
	# deskLocation = sheet.cell(row=row, column=3).value
	# sysId = sheet.cell(row=row, column=sheet.max_column).value
	for column in range(sheet.max_column,1,-1):
		variableUpdates[sheet.max_column - column]=sheet.cell(row=row, column=column).value
		

	newUrl = url + variableUpdates[0]
	
	dataUpdate='{"'# hardware_status":"' + str(hardwareStatus) + '", "install_status":"' + str(installStatus) + '", "u_desk_location":"' + str(deskLocation) + '"}'
	for column in range (1,sheet.max_column-2):
		dataUpdate = dataUpdate + str(variableNames[column]) + '":"' + str(variableUpdates[column]) + '","'
		
	dataUpdate = dataUpdate + str(variableNames[sheet.max_column-1]) + '":"' + str(variableUpdates[sheet.max_column-1]) + '"}'
	# Do the HTTP request
	response = requests.put(newUrl, auth=(user, pwd), headers=headers ,data=dataUpdate)

	# Check for HTTP codes other than 200
	if response.status_code != 200: 
		print('Status:', response.status_code, 'Headers:', response.headers, 'Error Response:',response.json())
		exit()
	 
	# Decode the JSON response into a dictionary and use the data
	print('Status:',response.status_code,'Headers:',response.headers,'Response:',response.json())